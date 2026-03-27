#!/usr/bin/env python3
# SPDX-FileCopyrightText: © 2026 Tenstorrent AI ULC
#
# SPDX-License-Identifier: Apache-2.0

"""LLK CodeGen Dashboard — local web UI for monitoring codegen runs."""

import json
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify

app = Flask(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
BASE_DIR = Path("/proj_sw/user_dev/llk_code_gen")

PROJECTS = {
    "quasar": {
        "name": "Quasar CodeGen",
        "desc": "Generate kernels for Quasar architecture",
        "logs_root": BASE_DIR / "quasar",
        "runs_jsonl": BASE_DIR / "quasar" / "runs.jsonl",
    },
    "blackhole": {
        "name": "Blackhole Issue Solver",
        "desc": "Debug and fix Blackhole kernel issues",
        "logs_root": BASE_DIR / "blackhole",
        "runs_jsonl": BASE_DIR / "blackhole" / "runs.jsonl",
    },
}


def get_project(project_id):
    return PROJECTS.get(project_id, PROJECTS["quasar"])


def load_runs(project_id="quasar"):
    """Load runs from runs.jsonl + scan run.json files in logs_root as fallback."""
    proj = get_project(project_id)
    runs_file = proj["runs_jsonl"]
    logs_root = proj["logs_root"]
    seen_ids = set()
    runs = []

    # 1. Load from runs.jsonl (primary source)
    if runs_file.exists():
        for line in runs_file.read_text().strip().splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                run = json.loads(line)
                seen_ids.add(run.get("run_id"))
                _enrich_run(run)
                runs.append(run)
            except (json.JSONDecodeError, ValueError):
                continue

    # 2. Scan run.json files in log directories (catch any not in JSONL)
    if logs_root.exists():
        for run_dir in sorted(logs_root.iterdir()):
            run_json = run_dir / "run.json"
            if not run_json.exists():
                continue
            try:
                run = json.loads(run_json.read_text())
                if run.get("run_id") in seen_ids:
                    continue
                seen_ids.add(run.get("run_id"))
                _enrich_run(run)
                runs.append(run)
            except (json.JSONDecodeError, ValueError):
                continue

    return runs


def _enrich_run(run):
    """Add derived display fields to a run dict."""
    try:
        start = datetime.fromisoformat(run.get("start_time", "").replace("Z", "+00:00"))
        end = datetime.fromisoformat(run.get("end_time", "").replace("Z", "+00:00"))
        run["duration_seconds"] = int((end - start).total_seconds())
        run["duration_display"] = format_duration(run["duration_seconds"])
    except (ValueError, TypeError):
        run["duration_seconds"] = 0
        run["duration_display"] = "—"
    run["display_status"] = compute_display_status(run)


def compute_display_status(run):
    """Compute three-way status: success / compiled / failed."""
    status = run.get("status", "failed")
    if status == "failed":
        return "failed"
    # It at least compiled — check tests
    tests_total = run.get("tests_total", 0)
    tests_passed = run.get("tests_passed", 0)
    if tests_total > 0 and tests_passed == tests_total:
        return "success"
    if status == "success" and tests_total > 0 and tests_passed == tests_total:
        return "success"
    return "compiled"


def format_duration(seconds):
    """Format seconds into human readable string."""
    if seconds < 60:
        return f"{seconds}s"
    minutes = seconds // 60
    secs = seconds % 60
    if minutes < 60:
        return f"{minutes}m {secs}s"
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}h {mins}m"


@app.route("/")
def index():
    return HTML_PAGE


@app.route("/api/projects")
def api_projects():
    return jsonify(
        [{"id": k, "name": v["name"], "desc": v["desc"]} for k, v in PROJECTS.items()]
    )


@app.route("/api/runs")
def api_runs():
    from flask import request

    project_id = request.args.get("project", "quasar")
    runs = load_runs(project_id)
    return jsonify(runs)


@app.route("/api/run/<run_id>")
def api_run_detail(run_id):
    from flask import request

    project_id = request.args.get("project", "quasar")
    proj = get_project(project_id)
    runs = load_runs(project_id)
    run = next((r for r in runs if r.get("run_id") == run_id), None)
    if not run:
        return jsonify({"error": "not found"}), 404

    # Load agent logs from LOG_DIR
    log_dir = proj["logs_root"] / run_id
    agent_logs = {}
    if log_dir.exists():
        for f in sorted(log_dir.iterdir()):
            if f.name.startswith("agent_") and f.suffix == ".md":
                agent_logs[f.stem] = f.read_text()

    # Load report
    report = None
    report_files = list(log_dir.glob("*_report.md")) if log_dir.exists() else []
    if report_files:
        report = report_files[0].read_text()

    # Load generated kernel source (try LOG_DIR snapshot first, fall back to repo)
    generated_code = None
    gen_file = run.get("generated_file", "")
    if gen_file:
        gen_filename = Path(gen_file).name
        snapshot_path = log_dir / gen_filename if log_dir.exists() else None
        if snapshot_path and snapshot_path.exists():
            generated_code = snapshot_path.read_text()
        else:
            gen_path = REPO_ROOT / gen_file
            if gen_path.exists():
                generated_code = gen_path.read_text()

    # Load reference kernel source (try LOG_DIR snapshot first, fall back to repo)
    reference_code = None
    ref_file = run.get("reference_file", "")
    if ref_file:
        ref_filename = Path(ref_file).name
        ref_snapshot = log_dir / f"ref_{ref_filename}" if log_dir.exists() else None
        if ref_snapshot and ref_snapshot.exists():
            reference_code = ref_snapshot.read_text()
        else:
            ref_path = REPO_ROOT / ref_file
            if ref_path.exists():
                reference_code = ref_path.read_text()

    # Load generated test files (if tests were generated by test-writer agent)
    test_cpp_code = None
    test_py_code = None
    test_cpp_path = None
    test_py_path = None
    if run.get("tests_generated"):
        kernel = run.get("kernel", "")
        arch = run.get("arch", "quasar")
        # C++ test source (try LOG_DIR snapshot first, fall back to repo)
        for pattern in [
            f"tests/sources/{arch}/{kernel}_{arch}_test.cpp",
            f"tests/sources/{arch}/sfpu_{kernel}_{arch}_test.cpp",
        ]:
            snapshot_name = Path(pattern).name
            snapshot = log_dir / snapshot_name if log_dir.exists() else None
            if snapshot and snapshot.exists():
                test_cpp_code = snapshot.read_text()
                test_cpp_path = pattern
                break
            p = REPO_ROOT / pattern
            if p.exists():
                test_cpp_code = p.read_text()
                test_cpp_path = pattern
                break
        # Python test driver (try LOG_DIR snapshot first, fall back to repo)
        for pattern in [
            f"tests/python_tests/{arch}/test_{kernel}_{arch}.py",
            f"tests/python_tests/{arch}/test_sfpu_{kernel}_{arch}.py",
        ]:
            snapshot_name = Path(pattern).name
            snapshot = log_dir / snapshot_name if log_dir.exists() else None
            if snapshot and snapshot.exists():
                test_py_code = snapshot.read_text()
                test_py_path = pattern
                break
            p = REPO_ROOT / pattern
            if p.exists():
                test_py_code = p.read_text()
                test_py_path = pattern
                break

    return jsonify(
        {
            "run": run,
            "agent_logs": agent_logs,
            "report": report,
            "generated_code": generated_code,
            "reference_code": reference_code,
            "test_cpp_code": test_cpp_code,
            "test_cpp_path": test_cpp_path,
            "test_py_code": test_py_code,
            "test_py_path": test_py_path,
            "log_dir": str(log_dir),
        }
    )


@app.route("/api/batches")
def api_batches():
    from flask import request

    project_id = request.args.get("project", "quasar")
    runs = load_runs(project_id)
    batches = {}
    for run in runs:
        bid = run.get("batch_id") or "manual"
        if bid not in batches:
            batches[bid] = {
                "batch_id": bid,
                "runs": [],
                "success": 0,
                "compiled": 0,
                "failed": 0,
                "total": 0,
            }
        batches[bid]["runs"].append(run["run_id"])
        batches[bid]["total"] += 1
        ds = run["display_status"]
        if ds in batches[bid]:
            batches[bid][ds] += 1
    return jsonify(list(batches.values()))


GAP_EXCEL = BASE_DIR / "quasar" / "quasar_llk_gap_analysis.xlsx"


def _categorize_llk(filename):
    """Categorize an llk_lib header by its prefix."""
    base = filename.split("/")[-1]  # strip experimental/ prefix if present
    if filename.startswith("experimental/"):
        return "Experimental"
    if base.startswith("llk_math_"):
        return "Math"
    elif base.startswith("llk_pack"):
        return "Pack"
    elif base.startswith("llk_unpack"):
        return "Unpack"
    return "Common"


# Known BH → QSR filename renames in llk_lib/
_LLK_RENAME_MAP = {
    "llk_unpack_A.h": "llk_unpack_unary_operand.h",
    "llk_unpack_AB.h": "llk_unpack_binary_operands.h",
    "llk_unpack_AB_matmul.h": "llk_unpack_matmul.h",
    "llk_unpack_AB_reduce.h": "llk_unpack_reduce.h",
    "llk_math_eltwise_unary_sfpu.h": "llk_math_eltwise_unary_sfpu_common.h",
}


def load_coverage():
    """Return coverage data — live file scan for SFPU & LLK Submodule, Excel for APIs."""
    result = {}

    # ── Live scan: SFPU Kernels ──────────────────────────────────
    bh_sfpu_dir = REPO_ROOT / "tt_llk_blackhole" / "common" / "inc" / "sfpu"
    qsr_sfpu_dir = REPO_ROOT / "tt_llk_quasar" / "common" / "inc" / "sfpu"

    bh_sfpu = (
        {
            f.stem.replace("ckernel_sfpu_", ""): f.name
            for f in sorted(bh_sfpu_dir.glob("ckernel_sfpu_*.h"))
        }
        if bh_sfpu_dir.exists()
        else {}
    )
    qsr_sfpu = (
        {
            f.stem.replace("ckernel_sfpu_", ""): f.name
            for f in sorted(qsr_sfpu_dir.glob("ckernel_sfpu_*.h"))
        }
        if qsr_sfpu_dir.exists()
        else {}
    )

    sfpu = []
    for name in sorted(bh_sfpu):
        in_qsr = name in qsr_sfpu
        sfpu.append(
            {
                "kernel": name,
                "bh": bh_sfpu[name],
                "qsr": qsr_sfpu[name] if in_qsr else "—",
                "status": "Implemented" if in_qsr else "MISSING",
            }
        )
    for name in sorted(qsr_sfpu):
        if name not in bh_sfpu:
            sfpu.append(
                {
                    "kernel": name,
                    "bh": "—",
                    "qsr": qsr_sfpu[name],
                    "status": "Quasar-specific",
                }
            )
    result["sfpu_kernels"] = sfpu

    # ── Live scan: LLK Submodule ─────────────────────────────────
    bh_llk_dir = REPO_ROOT / "tt_llk_blackhole" / "llk_lib"
    qsr_llk_dir = REPO_ROOT / "tt_llk_quasar" / "llk_lib"

    def _scan_llk_files(llk_dir):
        """Scan llk_lib/ including experimental/ subdirectory."""
        files = set()
        if not llk_dir.exists():
            return files
        for f in llk_dir.glob("*.h"):
            files.add(f.name)
        exp_dir = llk_dir / "experimental"
        if exp_dir.exists():
            for f in exp_dir.glob("*.h"):
                files.add(f"experimental/{f.name}")
        return files

    bh_llk = _scan_llk_files(bh_llk_dir)
    qsr_llk = _scan_llk_files(qsr_llk_dir)

    claimed_qsr = set()
    submodule = []

    for bh_file in sorted(bh_llk):
        qsr_file = None
        notes = ""
        if bh_file in qsr_llk:
            qsr_file = bh_file
        elif bh_file in _LLK_RENAME_MAP and _LLK_RENAME_MAP[bh_file] in qsr_llk:
            qsr_file = _LLK_RENAME_MAP[bh_file]
            notes = f"Renamed from {bh_file}"
        if qsr_file:
            claimed_qsr.add(qsr_file)
        submodule.append(
            {
                "category": _categorize_llk(bh_file),
                "bh_file": bh_file,
                "qsr_file": qsr_file or "MISSING",
                "status": "Implemented" if qsr_file else "MISSING",
                "notes": notes,
            }
        )

    for qsr_file in sorted(qsr_llk - claimed_qsr):
        submodule.append(
            {
                "category": _categorize_llk(qsr_file),
                "bh_file": "—",
                "qsr_file": qsr_file,
                "status": "Quasar-specific",
                "notes": "New in Quasar",
            }
        )
    result["submodule"] = submodule

    # ── Excel fallback: Intermediate API & Compute API ───────────
    try:
        import openpyxl
    except ImportError:
        result["intermediate_api"] = []
        result["compute_api"] = []
        result["summary"] = {}
        return result

    if not GAP_EXCEL.exists():
        result["intermediate_api"] = []
        result["compute_api"] = []
        result["summary"] = {}
        return result

    wb = openpyxl.load_workbook(str(GAP_EXCEL), data_only=True)

    # Sheet 2: Intermediate API
    ws = wb["2. Intermediate API"]
    api = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        api.append(
            {
                "bh_file": row[0],
                "qsr_file": row[1] or "MISSING",
                "status": row[2],
                "notes": row[3],
            }
        )
    result["intermediate_api"] = api

    # Sheet 1: Compute API (aggregate by file)
    ws = wb["1. Compute API"]
    compute_files = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        f = row[0]
        status = row[5] or "Unknown"
        if f not in compute_files:
            compute_files[f] = {
                "file": f,
                "total": 0,
                "implemented": 0,
                "stub": 0,
                "missing": 0,
            }
        compute_files[f]["total"] += 1
        if status == "Implemented":
            compute_files[f]["implemented"] += 1
        elif "stub" in (status or "").lower():
            compute_files[f]["stub"] += 1
        else:
            compute_files[f]["missing"] += 1
    result["compute_api"] = sorted(compute_files.values(), key=lambda x: -x["missing"])

    # Summary: computed dynamically from live + Excel data
    sfpu_impl = sum(1 for k in result["sfpu_kernels"] if k["status"] == "Implemented")
    sub_impl = sum(
        1
        for s in result["submodule"]
        if s["status"] in ("Implemented", "Quasar-specific")
    )
    api_impl = sum(
        1 for a in result["intermediate_api"] if a.get("status") == "Implemented"
    )
    compute_impl = sum(f["implemented"] for f in result["compute_api"])
    compute_total = sum(f["total"] for f in result["compute_api"])
    result["summary"] = {
        "SFPU Kernels": f"{sfpu_impl}/{len(result['sfpu_kernels'])}",
        "LLK Submodule": f"{sub_impl}/{len(result['submodule'])}",
        "Intermediate API": f"{api_impl}/{len(result['intermediate_api'])}",
        "Compute API": f"{compute_impl}/{compute_total}",
    }

    return result


@app.route("/api/coverage")
def api_coverage():
    return jsonify(load_coverage())


HTML_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>LLK CodeGen Dashboard</title>
<style>
  :root {
    --bg: #0d1117;
    --surface: #161b22;
    --border: #30363d;
    --text: #e6edf3;
    --text-muted: #8b949e;
    --green: #3fb950;
    --blue: #58a6ff;
    --red: #f85149;
    --yellow: #d29922;
    --green-bg: rgba(63,185,80,0.12);
    --blue-bg: rgba(88,166,255,0.12);
    --red-bg: rgba(248,81,73,0.12);
  }
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body { background: var(--bg); color: var(--text); font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif; font-size: 14px; }

  .header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 16px 24px; display: flex; align-items: center; justify-content: space-between; }
  .header h1 { font-size: 20px; font-weight: 600; }
  .header .stats { display: flex; gap: 16px; }

  .tabs { display: flex; gap: 0; border-bottom: 1px solid var(--border); background: var(--surface); padding: 0 24px; }
  .tab { padding: 10px 16px; cursor: pointer; color: var(--text-muted); border-bottom: 2px solid transparent; font-size: 14px; }
  .tab:hover { color: var(--text); }
  .tab.active { color: var(--text); border-bottom-color: var(--blue); }

  .content { padding: 24px; max-width: 1400px; margin: 0 auto; }

  /* Summary cards */
  .summary-cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 24px; }
  .card { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 16px; }
  .card .label { color: var(--text-muted); font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px; }
  .card .value { font-size: 28px; font-weight: 600; }
  .card .value.green { color: var(--green); }
  .card .value.blue { color: var(--blue); }
  .card .value.red { color: var(--red); }

  /* Filter bar */
  .filter-bar { display: flex; gap: 12px; margin-bottom: 16px; align-items: center; }
  .filter-bar select, .filter-bar input { background: var(--surface); color: var(--text); border: 1px solid var(--border); border-radius: 6px; padding: 6px 12px; font-size: 13px; }
  .filter-bar label { color: var(--text-muted); font-size: 13px; }

  /* Runs table */
  .runs-table { width: 100%; border-collapse: collapse; }
  .runs-table th { text-align: left; padding: 8px 12px; color: var(--text-muted); font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; border-bottom: 1px solid var(--border); }
  .runs-table td { padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: top; }
  .runs-table tr:hover { background: rgba(88,166,255,0.04); cursor: pointer; }
  .runs-table tr:last-child td { border-bottom: none; }

  .status-badge { display: inline-flex; align-items: center; gap: 6px; padding: 2px 10px; border-radius: 12px; font-size: 12px; font-weight: 600; text-transform: uppercase; }
  .status-badge.success { background: var(--green-bg); color: var(--green); }
  .status-badge.compiled { background: var(--blue-bg); color: var(--blue); }
  .status-badge.failed { background: var(--red-bg); color: var(--red); }
  .status-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; }
  .status-dot.success { background: var(--green); }
  .status-dot.compiled { background: var(--blue); }
  .status-dot.failed { background: var(--red); }

  .obstacle-text { color: var(--text-muted); font-size: 12px; margin-top: 4px; }
  .kernel-type { color: var(--text-muted); font-size: 11px; }
  .run-type-badge { display: inline-block; padding: 1px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; text-transform: uppercase; }
  .run-type-badge.ci { background: #1a3a4a; color: #4fc3f7; }
  .run-type-badge.manual { background: #3a3a2a; color: #c0b060; }

  /* Detail panel */
  .detail-panel { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 24px; }
  .detail-panel h2 { margin-bottom: 16px; font-size: 18px; }
  .detail-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px; }
  .detail-field .label { color: var(--text-muted); font-size: 12px; margin-bottom: 2px; }
  .detail-field .value { font-size: 14px; }

  .phase-card { background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 12px 16px; margin-bottom: 8px; }
  .phase-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; }
  .phase-name { font-weight: 600; }
  .phase-stats { display: flex; gap: 16px; font-size: 13px; color: var(--text-muted); }
  .phase-error { color: var(--red); font-size: 12px; margin-top: 4px; font-family: monospace; }
  .phase-test-detail { color: var(--text-muted); font-size: 12px; margin-top: 4px; }

  /* Failures */
  .failure-item { background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 12px 16px; margin-bottom: 8px; }
  .failure-item.unresolved { border-left: 3px solid var(--red); }
  .failure-item.resolved { border-left: 3px solid var(--green); }
  .failure-tag { display: inline-block; padding: 1px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; text-transform: uppercase; }
  .failure-tag.compile_error { background: #3a2a1a; color: #f0a050; }
  .failure-tag.test_failure { background: var(--red-bg); color: var(--red); }
  .failure-tag.agent_error { background: #3a2a3a; color: #c070c0; }
  .failure-tag.infra_error { background: #1a3a4a; color: #4fc3f7; }
  .failure-tag.resolved { background: var(--green-bg); color: var(--green); }
  .failure-tag.unresolved { background: var(--red-bg); color: var(--red); }

  .agent-logs { margin-top: 16px; }
  .agent-log-btn { background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 6px 12px; color: var(--blue); cursor: pointer; font-size: 13px; margin-right: 8px; margin-bottom: 8px; display: inline-block; }
  .agent-log-btn:hover { border-color: var(--blue); }
  .agent-log-btn.active { border-color: var(--blue); background: var(--blue-bg); }
  .log-content { background: var(--bg); border: 1px solid var(--border); border-radius: 6px; padding: 16px; margin-top: 12px; white-space: pre-wrap; font-family: monospace; font-size: 12px; max-height: 500px; overflow-y: auto; line-height: 1.6; }

  .back-btn { background: none; border: 1px solid var(--border); border-radius: 6px; padding: 6px 16px; color: var(--text); cursor: pointer; margin-bottom: 16px; font-size: 13px; }
  .back-btn:hover { border-color: var(--blue); color: var(--blue); }

  /* Pipeline */
  .pipeline { display: flex; align-items: flex-start; gap: 0; margin-bottom: 24px; overflow-x: auto; padding: 8px 0; }
  .pipeline-step { display: flex; align-items: center; }
  .pipeline-node { background: var(--surface); border: 2px solid var(--border); border-radius: 8px; padding: 10px 14px; min-width: 120px; text-align: center; cursor: default; transition: border-color 0.2s; }
  .pipeline-node:hover { border-color: var(--blue); }
  .pipeline-node.ran { border-color: var(--green); }
  .pipeline-node.skipped { border-color: var(--border); opacity: 0.45; }
  .pipeline-node .step-num { font-size: 10px; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.5px; }
  .pipeline-node .step-name { font-size: 13px; font-weight: 600; margin: 4px 0 2px; }
  .pipeline-node .step-desc { font-size: 11px; color: var(--text-muted); line-height: 1.3; }
  .pipeline-node .step-status { font-size: 11px; margin-top: 4px; }
  .pipeline-arrow { color: var(--border); font-size: 18px; padding: 0 6px; margin-top: 14px; flex-shrink: 0; }
  .pipeline-node.ran .step-status { color: var(--green); }
  .pipeline-loop { display: flex; flex-direction: column; align-items: center; }
  .pipeline-loop-label { font-size: 10px; color: var(--yellow); margin-bottom: 4px; }

  /* Trends */
  .trend-section { margin-bottom: 32px; }
  .trend-section h3 { margin-bottom: 12px; color: var(--text-muted); font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; }

  .heatmap { width: 100%; border-collapse: collapse; }
  .heatmap th { text-align: left; padding: 6px 10px; color: var(--text-muted); font-size: 12px; }
  .heatmap td { padding: 6px 10px; text-align: center; }
  .heatmap-cell { width: 24px; height: 24px; border-radius: 4px; display: inline-block; }
  .heatmap-cell.success { background: var(--green); }
  .heatmap-cell.compiled { background: var(--blue); }
  .heatmap-cell.failed { background: var(--red); }
  .heatmap-cell.empty { background: var(--border); opacity: 0.3; }

  .chart-container { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 16px; margin-bottom: 16px; }
  .bar-chart { display: flex; align-items: flex-end; gap: 4px; height: 120px; padding: 0 8px; }
  .bar-group { display: flex; flex-direction: column; align-items: center; flex: 1; height: 100%; justify-content: flex-end; }
  .bar { width: 100%; max-width: 40px; border-radius: 3px 3px 0 0; min-height: 2px; transition: height 0.3s; }
  .bar.success { background: var(--green); }
  .bar.compiled { background: var(--blue); }
  .bar.failed { background: var(--red); }
  .bar-label { font-size: 10px; color: var(--text-muted); margin-top: 4px; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 60px; }
  .bar-value { font-size: 10px; color: var(--text-muted); margin-bottom: 2px; }

  .empty-state { text-align: center; padding: 60px 20px; color: var(--text-muted); }
  .empty-state h2 { margin-bottom: 8px; color: var(--text); }

  /* Responsive */
  @media (max-width: 768px) {
    .detail-grid { grid-template-columns: 1fr; }
    .summary-cards { grid-template-columns: repeat(2, 1fr); }
  }
</style>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/styles/github-dark.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/highlight.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/cpp.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.9.0/languages/python.min.js"></script>
<style>
  .hljs { background: var(--bg) !important; padding: 0 !important; }
  .code-block { font-size: 13px; line-height: 1.5; overflow-x: auto; }
  .code-block table { border-collapse: collapse; width: 100%; }
  .code-block .line-num { color: var(--text-muted); user-select: none; text-align: right; padding-right: 12px; white-space: nowrap; vertical-align: top; }
  .code-block .line-code { white-space: pre; font-family: monospace; }
</style>
</head>
<body>

<div class="header">
  <div style="display:flex;align-items:center;gap:16px">
    <h1>LLK CodeGen Dashboard</h1>
    <select id="project-select" onchange="switchProject(this.value)" style="background:var(--bg);color:var(--text);border:1px solid var(--border);border-radius:6px;padding:6px 12px;font-size:13px;cursor:pointer"></select>
  </div>
  <div class="stats" id="header-stats"></div>
</div>

<div class="tabs">
  <div class="tab active" data-tab="overview" onclick="switchTab('overview')">Overview</div>
  <div class="tab" data-tab="trends" onclick="switchTab('trends')">Trends</div>
  <div class="tab" data-tab="coverage" onclick="switchTab('coverage')">Coverage</div>
  <div class="tab" data-tab="detail" id="detail-tab" onclick="switchTab('detail')" style="display:none">Run Detail</div>
</div>

<div class="content" id="content"></div>

<script>
let allRuns = [];
let currentTab = 'overview';
let currentRunId = null;
let currentDetail = null;
let currentProject = 'quasar';
let projects = [];
let filters = { batch: 'all', type: 'all', status: 'all', date: '', model: 'all', runtype: 'all' };

async function init() {
  // Load projects
  const projResp = await fetch('/api/projects');
  projects = await projResp.json();
  const projSelect = document.getElementById('project-select');
  projSelect.innerHTML = projects.map(p =>
    `<option value="${p.id}" ${p.id === currentProject ? 'selected' : ''}>${p.name}</option>`
  ).join('');

  await loadRuns();
}

async function loadRuns() {
  const resp = await fetch('/api/runs?project=' + currentProject);
  allRuns = await resp.json();
  // Sort newest first
  allRuns.sort((a, b) => (b.start_time || '').localeCompare(a.start_time || ''));
  updateHeaderStats();
  if (currentTab === 'detail' && currentRunId) renderDetail();
  else if (currentTab === 'trends') renderTrends();
  else renderOverview();
}

async function switchProject(projectId) {
  currentProject = projectId;
  currentRunId = null;
  currentDetail = null;
  document.getElementById('detail-tab').style.display = 'none';
  filters = { batch: 'all', type: 'all', status: 'all', date: '', model: 'all', runtype: 'all' };
  coverageData = null;
  switchTab('overview');
  await loadRuns();
}

function updateHeaderStats() {
  const el = document.getElementById('header-stats');
  const total = allRuns.length;
  const success = allRuns.filter(r => r.display_status === 'success').length;
  const compiled = allRuns.filter(r => r.display_status === 'compiled').length;
  const failed = allRuns.filter(r => r.display_status === 'failed').length;
  el.innerHTML = `
    <span style="color:var(--green)">${success} passed</span>
    <span style="color:var(--blue)">${compiled} compiled</span>
    <span style="color:var(--red)">${failed} failed</span>
    <span style="color:var(--text-muted)">${total} total</span>
  `;
}

function switchTab(tab) {
  currentTab = tab;
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelector(`.tab[data-tab="${tab}"]`).classList.add('active');
  if (tab === 'overview') renderOverview();
  else if (tab === 'trends') renderTrends();
  else if (tab === 'coverage') renderCoverage();
  else if (tab === 'detail') renderDetail();
}

// ==================== OVERVIEW ====================

function saveFilters() {
  const b = document.getElementById('filter-batch');
  const t = document.getElementById('filter-type');
  const s = document.getElementById('filter-status');
  const d = document.getElementById('filter-date');
  const m = document.getElementById('filter-model');
  const rt = document.getElementById('filter-runtype');
  if (b) filters.batch = b.value;
  if (t) filters.type = t.value;
  if (s) filters.status = s.value;
  if (d) filters.date = d.value;
  if (m) filters.model = m.value;
  if (rt) filters.runtype = rt.value;
}

function sel(current, value) { return current === value ? 'selected' : ''; }

function renderOverview() {
  saveFilters();
  const content = document.getElementById('content');
  const runs = getFilteredRuns();

  const success = runs.filter(r => r.display_status === 'success').length;
  const compiled = runs.filter(r => r.display_status === 'compiled').length;
  const failed = runs.filter(r => r.display_status === 'failed').length;
  const totalCompiles = runs.reduce((s, r) => s + (r.compilation_attempts || 0), 0);
  const totalPhases = runs.reduce((s, r) => s + (r.phases_total || 0), 0);
  const avgCompiles = totalPhases > 0 ? (totalCompiles / totalPhases).toFixed(1) : '—';
  const totalTokens = runs.reduce((s, r) => s + ((r.tokens && r.tokens.total) || 0), 0);
  const totalDuration = runs.reduce((s, r) => s + (r.duration_seconds || 0), 0);

  // Get unique batches for filter
  const batches = [...new Set(allRuns.map(r => r.batch_id || 'manual'))];

  content.innerHTML = `
    <div class="filter-bar">
      <label>Batch:</label>
      <select id="filter-batch" onchange="renderOverview()">
        <option value="all" ${sel(filters.batch,'all')}>All runs</option>
        ${batches.map(b => `<option value="${b}" ${sel(filters.batch,b)}>${b}</option>`).join('')}
      </select>
      <label>Kernel type:</label>
      <select id="filter-type" onchange="renderOverview()">
        <option value="all" ${sel(filters.type,'all')}>All</option>
        <option value="sfpu" ${sel(filters.type,'sfpu')}>sfpu</option>
        <option value="math" ${sel(filters.type,'math')}>math</option>
        <option value="pack" ${sel(filters.type,'pack')}>pack</option>
        <option value="unpack" ${sel(filters.type,'unpack')}>unpack</option>
      </select>
      <label>Status:</label>
      <select id="filter-status" onchange="renderOverview()">
        <option value="all" ${sel(filters.status,'all')}>All</option>
        <option value="success" ${sel(filters.status,'success')}>Success</option>
        <option value="compiled" ${sel(filters.status,'compiled')}>Compiled</option>
        <option value="failed" ${sel(filters.status,'failed')}>Failed</option>
      </select>
      <label>Model:</label>
      <select id="filter-model" onchange="renderOverview()">
        <option value="all" ${sel(filters.model,'all')}>All</option>
        ${[...new Set(allRuns.map(r => r.model || 'unknown').filter(m => m !== 'unknown'))].sort().map(m => `<option value="${m}" ${sel(filters.model,m)}>${m}</option>`).join('')}
      </select>
      <label>Run type:</label>
      <select id="filter-runtype" onchange="renderOverview()">
        <option value="all" ${sel(filters.runtype,'all')}>All</option>
        <option value="ci" ${sel(filters.runtype,'ci')}>CI</option>
        <option value="manual" ${sel(filters.runtype,'manual')}>Manual</option>
      </select>
      <label>Date:</label>
      <input type="date" id="filter-date" value="${filters.date}" onchange="renderOverview()">
    </div>

    <div class="summary-cards">
      <div class="card">
        <div class="label">Success</div>
        <div class="value green">${success}</div>
      </div>
      <div class="card">
        <div class="label">Compiled</div>
        <div class="value blue">${compiled}</div>
      </div>
      <div class="card">
        <div class="label">Failed</div>
        <div class="value red">${failed}</div>
      </div>
      <div class="card">
        <div class="label">Avg Compiles/Phase</div>
        <div class="value">${avgCompiles}</div>
      </div>
      <div class="card">
        <div class="label">Total Tokens</div>
        <div class="value">${formatTokens(totalTokens)}</div>
      </div>
      <div class="card">
        <div class="label">Total Time</div>
        <div class="value">${formatDuration(totalDuration)}</div>
      </div>
    </div>

    ${runs.length === 0 ? `
      <div class="empty-state">
        <h2>No runs yet</h2>
        <p>Run a codegen prompt and results will appear here.</p>
      </div>
    ` : `
      <table class="runs-table">
        <thead>
          <tr>
            <th>Status</th>
            <th>Kernel</th>
            <th>Prompt</th>
            <th>Model</th>
            <th>Type</th>
            <th>Phases</th>
            <th>Compiles</th>
            <th>Debug</th>
            <th>Tests</th>
            <th>Duration</th>
            <th>Date</th>
          </tr>
        </thead>
        <tbody>
          ${runs.map(r => renderRunRow(r)).join('')}
        </tbody>
      </table>
    `}
  `;
}

function getFilteredRuns() {
  let runs = [...allRuns];
  if (filters.batch !== 'all') {
    runs = runs.filter(r => (r.batch_id || 'manual') === filters.batch);
  }
  if (filters.type !== 'all') {
    runs = runs.filter(r => r.kernel_type === filters.type);
  }
  if (filters.status !== 'all') {
    runs = runs.filter(r => r.display_status === filters.status);
  }
  if (filters.model !== 'all') {
    runs = runs.filter(r => (r.model || 'unknown') === filters.model);
  }
  if (filters.runtype !== 'all') {
    runs = runs.filter(r => (r.run_type || 'manual') === filters.runtype);
  }
  if (filters.date) {
    runs = runs.filter(r => (r.start_time || '').slice(0, 10) === filters.date);
  }
  return runs;
}

function renderRunRow(r) {
  const ds = r.display_status;
  const date = r.start_time ? r.start_time.slice(0, 10) : '—';
  const tests = r.tests_total > 0 ? `${r.tests_passed}/${r.tests_total}` : '—';
  const testTag = r.tests_generated ? ' <span style="font-size:10px;color:var(--blue);cursor:help" title="No pre-existing test was found for this kernel, so the test-writer agent generated tests as part of the codegen run">AI generated</span>' : r.tests_total > 0 ? ' <span style="font-size:10px;color:var(--text-muted);cursor:help" title="Used pre-existing tests that were already in the repo before this run">existing</span>' : '';
  return `
    <tr onclick="openDetail('${r.run_id}')">
      <td><span class="status-badge ${ds}"><span class="status-dot ${ds}"></span>${ds}</span></td>
      <td><strong>${r.kernel}</strong><br><span class="kernel-type">${r.kernel_type} · ${r.arch}</span></td>
      <td style="max-width:250px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.prompt || '—'}</td>
      <td>${r.model || '—'}</td>
      <td><span class="run-type-badge ${r.run_type || 'manual'}">${r.run_type || 'manual'}</span></td>
      <td>${r.phases_completed || 0}/${r.phases_total || 0}</td>
      <td>${r.compilation_attempts || 0}</td>
      <td>${r.debug_cycles || 0}</td>
      <td>${tests}${testTag}</td>
      <td>${r.duration_display || '—'}</td>
      <td>${date}</td>
    </tr>
    ${r.obstacle ? `<tr onclick="openDetail('${r.run_id}')"><td></td><td colspan="10" class="obstacle-text">⚠ ${r.obstacle}</td></tr>` : ''}
  `;
}

// ==================== DETAIL ====================

async function openDetail(runId) {
  currentRunId = runId;
  document.getElementById('detail-tab').style.display = '';
  switchTab('detail');
}

async function renderDetail() {
  const content = document.getElementById('content');
  if (!currentRunId) {
    content.innerHTML = '<div class="empty-state"><p>Select a run from Overview.</p></div>';
    return;
  }

  content.innerHTML = '<div class="empty-state"><p>Loading...</p></div>';

  const resp = await fetch(`/api/run/${currentRunId}?project=${currentProject}`);
  const data = await resp.json();
  currentDetail = data;
  const r = data.run;
  const ds = r.display_status;
  const tests = r.tests_total > 0 ? `${r.tests_passed}/${r.tests_total}` : 'N/A';

  content.innerHTML = `
    <button class="back-btn" onclick="switchTab('overview')">← Back to Overview</button>

    <div class="detail-panel">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
        <h2>${r.kernel} <span class="kernel-type" style="font-size:14px;font-weight:normal">${r.kernel_type} · ${r.arch}</span></h2>
        <span class="status-badge ${ds}"><span class="status-dot ${ds}"></span>${ds}</span>
      </div>

      <div class="detail-grid">
        <div class="detail-field"><div class="label">Prompt</div><div class="value">${r.prompt || '—'}</div></div>
        <div class="detail-field"><div class="label">Run ID</div><div class="value" style="font-family:monospace;font-size:12px">${r.run_id}</div></div>
        <div class="detail-field"><div class="label">Duration</div><div class="value">${r.duration_display}</div></div>
        <div class="detail-field"><div class="label">Lines Generated</div><div class="value">${r.lines_generated}</div></div>
        <div class="detail-field"><div class="label">Reference</div><div class="value" style="font-size:12px">${r.reference_file}</div></div>
        <div class="detail-field"><div class="label">Generated</div><div class="value" style="font-size:12px">${r.generated_file}</div></div>
        <div class="detail-field"><div class="label">Compile Attempts</div><div class="value">${r.compilation_attempts}</div></div>
        <div class="detail-field"><div class="label">Debug Cycles</div><div class="value">${r.debug_cycles}</div></div>
        <div class="detail-field"><div class="label">Tests</div><div class="value">${tests}</div></div>
        <div class="detail-field"><div class="label">Test Source</div><div class="value">${r.tests_generated ? '<span style="color:var(--blue)">Generated (test-writer agent)</span>' : r.tests_total > 0 ? 'Pre-existing' : 'No tests run'}</div></div>
        <div class="detail-field"><div class="label">Prettified</div><div class="value">${r.prettified ? 'Yes' : 'No'}</div></div>
        <div class="detail-field"><div class="label">Tokens</div><div class="value">${r.tokens && r.tokens.total > 0 ? formatTokens(r.tokens.total) + ` (in: ${formatTokens(r.tokens.input)}, out: ${formatTokens(r.tokens.output)}, cache: ${formatTokens(r.tokens.cache_read)})` : 'Not captured'}</div></div>
        <div class="detail-field"><div class="label">Model</div><div class="value">${r.model || 'Unknown'}</div></div>
        <div class="detail-field"><div class="label">Run Type</div><div class="value"><span class="run-type-badge ${r.run_type || 'manual'}">${r.run_type === 'ci' ? 'CI (scheduled)' : 'Manual'}</span></div></div>
        <div class="detail-field"><div class="label">Batch</div><div class="value">${r.batch_id || 'Manual run'}</div></div>
        <div class="detail-field"><div class="label">Git Commit</div><div class="value" style="font-family:monospace;font-size:12px">${r.git_commit || 'Not captured'}</div></div>
      </div>

      ${r.obstacle ? `<div style="background:var(--red-bg);border:1px solid var(--red);border-radius:6px;padding:12px;margin-bottom:16px;color:var(--red)">⚠ Obstacle: ${r.obstacle}</div>` : ''}

      <h3 style="margin-bottom:12px">Pipeline</h3>
      <div class="pipeline">
        ${renderPipeline(r, data.agent_logs)}
      </div>

      <h3 style="margin-bottom:12px">Phases</h3>
      ${(r.per_phase || []).map(p => `
        <div class="phase-card">
          <div class="phase-header">
            <span class="phase-name">Phase ${p.phase}: ${p.name}</span>
          </div>
          <div class="phase-stats">
            <span style="color: ${p.compilation_attempts <= 1 ? 'var(--green)' : 'var(--yellow)'}">
              ${p.compilation_attempts <= 1 ? 'Compiled on first try' : 'Compiled after ' + p.compilation_attempts + ' attempts'}
            </span>
            ${p.debug_cycles > 0 ? `<span style="color:var(--yellow)">${p.debug_cycles} debug cycle${p.debug_cycles > 1 ? 's' : ''}</span>` : ''}
          </div>
          ${p.compile_errors && p.compile_errors.length > 0 ? `
            <div style="margin-top:8px">
              <div style="font-size:12px;color:var(--text-muted);margin-bottom:4px">Compilation errors (${p.compile_errors.length}):</div>
              ${p.compile_errors.map(e => `
                <div class="phase-error" style="margin-bottom:4px">
                  <span style="color:var(--text-muted);font-size:11px">attempt ${e.attempt}:</span> ${e.error}
                </div>
              `).join('')}
            </div>
          ` : p.first_compile_error ? `<div class="phase-error">Initial error (fixed): ${p.first_compile_error}</div>` : ''}
          <div style="margin-top:8px;font-size:13px">
            ${p.test_result === 'passed' ? '<span style="color:var(--green)">Tests passed</span>' + (p.test_details ? ' — ' + p.test_details : '') :
              p.test_result === 'failed' ? '<span style="color:var(--red)">Tests failed</span>' + (p.test_details ? ' — ' + p.test_details : '') :
              '<span style="color:var(--text-muted)">Tests not run</span>' + (p.test_details ? ' — ' + p.test_details : '')}
          </div>
        </div>
      `).join('')}

      ${(r.failures && r.failures.length > 0) ? `
        <h3 style="margin-bottom:12px">Failures (${r.failures.length})</h3>
        ${r.failures.map(f => `
          <div class="failure-item ${f.resolved ? 'resolved' : 'unresolved'}">
            <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:6px">
              <span class="failure-tag ${f.type}">${(f.type || '').replace('_', ' ')}</span>
              <span class="failure-tag ${f.resolved ? 'resolved' : 'unresolved'}">${f.resolved ? 'resolved' : 'unresolved'}</span>
              <span style="font-size:12px;color:var(--text-muted)">step: <strong>${f.step || '—'}</strong> · agent: <strong>${f.agent || '—'}</strong></span>
            </div>
            <div style="font-family:monospace;font-size:12px;color:var(--text);white-space:pre-wrap;word-break:break-word;line-height:1.5">${escapeHtml(f.message || '')}</div>
          </div>
        `).join('')}
      ` : ''}

      <div class="agent-logs">
        <h3 style="margin-bottom:12px">Code</h3>
        <div id="code-btns">
          ${data.generated_code ? '<span class="agent-log-btn" onclick="showCode(&quot;generated&quot;, this)">Generated Kernel</span>' : ''}
          ${data.reference_code ? '<span class="agent-log-btn" onclick="showCode(&quot;reference&quot;, this)">Reference Kernel</span>' : ''}
          ${data.generated_code && data.reference_code ? '<span class="agent-log-btn" onclick="showCode(&quot;sidebyside&quot;, this)">Side by Side</span>' : ''}
          ${data.test_cpp_code ? '<span class="agent-log-btn" onclick="showCode(&quot;test_cpp&quot;, this)" style="border-color:var(--blue)">Generated Test (C++)</span>' : ''}
          ${data.test_py_code ? '<span class="agent-log-btn" onclick="showCode(&quot;test_py&quot;, this)" style="border-color:var(--blue)">Generated Test (Python)</span>' : ''}
          ${!data.generated_code && !data.reference_code ? '<span style="color:var(--text-muted)">No kernel files available</span>' : ''}
        </div>
        <div id="code-content"></div>
      </div>

      <div class="agent-logs" style="margin-top:24px">
        <h3 style="margin-bottom:12px">Agent Logs</h3>
        <div id="agent-btns">
          ${orderAgentLogs(Object.keys(data.agent_logs || {})).map(name =>
            `<span class="agent-log-btn" onclick="showAgentLog('${name}', this)">${name.replace('agent_', '')}</span>`
          ).join('')}
          ${Object.keys(data.agent_logs || {}).length === 0 ? '<span style="color:var(--text-muted)">No agent logs available</span>' : ''}
        </div>
        <div id="agent-log-content"></div>
      </div>
    </div>
  `;
}

function formatCodeBlock(code, path, lang) {
  lang = lang || 'cpp';
  const highlighted = hljs.highlight(code, {language: lang}).value;
  const lines = highlighted.split('\\n');
  const rows = lines.map((line, i) =>
    `<tr><td class="line-num">${i+1}</td><td class="line-code">${line}</td></tr>`
  ).join('');
  return `<div style="color:var(--text-muted);font-size:11px;margin-bottom:4px">${path}</div>` +
    `<div class="log-content code-block"><table>${rows}</table></div>`;
}

function showCode(which, btn) {
  const el = document.getElementById('code-content');
  if (btn.classList.contains('active')) {
    btn.classList.remove('active');
    el.innerHTML = '';
    return;
  }
  document.querySelectorAll('#code-btns .agent-log-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');

  if (which === 'sidebyside') {
    const gen = currentDetail.generated_code;
    const ref = currentDetail.reference_code;
    if (!gen || !ref) { el.innerHTML = '<p style="color:var(--text-muted);margin-top:12px">Both files needed for side-by-side view.</p>'; return; }
    el.innerHTML = `
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px">
        <div>
          <div style="font-weight:600;margin-bottom:6px;color:var(--green)">Generated (Quasar)</div>
          ${formatCodeBlock(gen, currentDetail.run.generated_file)}
        </div>
        <div>
          <div style="font-weight:600;margin-bottom:6px;color:var(--blue)">Reference (Blackhole)</div>
          ${formatCodeBlock(ref, currentDetail.run.reference_file)}
        </div>
      </div>`;
    return;
  }

  const codeMap = {
    generated: { code: currentDetail.generated_code, path: currentDetail.run.generated_file, lang: 'cpp' },
    reference: { code: currentDetail.reference_code, path: currentDetail.run.reference_file, lang: 'cpp' },
    test_cpp:  { code: currentDetail.test_cpp_code, path: currentDetail.test_cpp_path, lang: 'cpp' },
    test_py:   { code: currentDetail.test_py_code, path: currentDetail.test_py_path, lang: 'python' },
  };
  const entry = codeMap[which];
  if (!entry || !entry.code) { el.innerHTML = '<p style="color:var(--text-muted);margin-top:12px">File not found.</p>'; return; }
  el.innerHTML = `<div style="margin-top:12px">${formatCodeBlock(entry.code, entry.path, entry.lang)}</div>`;
}

function showAgentLog(name, btn) {
  const logEl = document.getElementById('agent-log-content');
  if (btn.classList.contains('active')) {
    btn.classList.remove('active');
    logEl.innerHTML = '';
    return;
  }
  document.querySelectorAll('#agent-btns .agent-log-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  const logs = currentDetail.agent_logs || {};
  logEl.innerHTML = logs[name] ? `<div class="log-content">${escapeHtml(logs[name])}</div>` : '<p style="color:var(--text-muted);margin-top:12px">No log content.</p>';
}

// ==================== TRENDS ====================

function renderTrends() {
  const content = document.getElementById('content');
  if (allRuns.length === 0) {
    content.innerHTML = '<div class="empty-state"><h2>No data yet</h2><p>Run some codegen prompts to see trends.</p></div>';
    return;
  }

  // Group by week
  const weeks = {};
  allRuns.forEach(r => {
    const d = r.start_time ? r.start_time.slice(0, 10) : 'unknown';
    // Group by ISO week
    const dt = new Date(r.start_time);
    const weekStart = new Date(dt);
    weekStart.setDate(dt.getDate() - dt.getDay());
    const wk = weekStart.toISOString().slice(0, 10);
    if (!weeks[wk]) weeks[wk] = [];
    weeks[wk].push(r);
  });
  const sortedWeeks = Object.keys(weeks).sort();

  // Get all unique kernels
  const allKernels = [...new Set(allRuns.map(r => r.kernel))].sort();

  // Compile attempts bar chart data
  const compileData = sortedWeeks.map(w => {
    const runs = weeks[w];
    const totalCompiles = runs.reduce((s, r) => s + (r.compilation_attempts || 0), 0);
    const totalPhases = runs.reduce((s, r) => s + (r.phases_total || 0), 0);
    return { week: w, avg: totalPhases > 0 ? totalCompiles / totalPhases : 0 };
  });
  const maxAvg = Math.max(...compileData.map(d => d.avg), 1);

  content.innerHTML = `
    <div class="trend-section">
      <h3>Weekly CI Success Rate</h3>
      ${(() => {
        const ciRuns = allRuns.filter(r => (r.run_type || 'manual') === 'ci');
        if (ciRuns.length === 0) return '<p style="color:var(--text-muted)">No CI runs yet. Run ci_weekly.py to start tracking.</p>';
        const ciWeeks = {};
        ciRuns.forEach(r => {
          const dt = new Date(r.start_time);
          const weekStart = new Date(dt);
          weekStart.setDate(dt.getDate() - dt.getDay());
          const wk = weekStart.toISOString().slice(0, 10);
          if (!ciWeeks[wk]) ciWeeks[wk] = [];
          ciWeeks[wk].push(r);
        });
        const sortedCiWeeks = Object.keys(ciWeeks).sort();
        return '<div class="chart-container"><div class="bar-chart">' +
          sortedCiWeeks.map((w, i) => {
            const wr = ciWeeks[w];
            const s = wr.filter(r => r.display_status === 'success').length;
            const pct = Math.round(s / wr.length * 100);
            const prev = i > 0 ? (() => { const pw = ciWeeks[sortedCiWeeks[i-1]]; return Math.round(pw.filter(r => r.display_status === 'success').length / pw.length * 100); })() : null;
            const delta = prev !== null ? pct - prev : null;
            const color = pct >= 80 ? 'var(--green)' : pct >= 50 ? 'var(--yellow)' : 'var(--red)';
            const deltaStr = delta !== null ? (delta > 0 ? '<span style="color:var(--green)">▲' + delta + '</span>' : delta < 0 ? '<span style="color:var(--red)">▼' + Math.abs(delta) + '</span>' : '<span style="color:var(--text-muted)">—</span>') : '';
            return '<div class="bar-group">' +
              '<div class="bar-value" style="color:' + color + '">' + pct + '%</div>' +
              '<div class="bar" style="height:' + Math.max(8, pct) + '%;background:' + color + '"></div>' +
              '<div class="bar-label">' + w.slice(5) + '</div>' +
              '<div style="font-size:10px;color:var(--text-muted)">' + s + '/' + wr.length + '</div>' +
              '<div style="font-size:10px">' + deltaStr + '</div>' +
            '</div>';
          }).join('') +
        '</div></div>';
      })()}
    </div>

    <div class="trend-section">
      <h3>Status Distribution by Run</h3>
      <div class="chart-container">
        <div class="bar-chart">
          ${allRuns.slice().reverse().map(r => `
            <div class="bar-group">
              <div class="bar-value">${r.compilation_attempts || 0}c</div>
              <div class="bar ${r.display_status}" style="height:${Math.max(8, (r.compilation_attempts || 1) / Math.max(...allRuns.map(x => x.compilation_attempts || 1)) * 100)}%"></div>
              <div class="bar-label" title="${r.kernel}">${r.kernel}</div>
            </div>
          `).join('')}
        </div>
      </div>
    </div>

    <div class="trend-section">
      <h3>Avg Compile Attempts per Phase (by Week)</h3>
      <div class="chart-container">
        <div class="bar-chart">
          ${compileData.map(d => `
            <div class="bar-group">
              <div class="bar-value">${d.avg.toFixed(1)}</div>
              <div class="bar compiled" style="height:${Math.max(8, d.avg / maxAvg * 100)}%"></div>
              <div class="bar-label">${d.week.slice(5)}</div>
            </div>
          `).join('')}
        </div>
      </div>
    </div>

    <div class="trend-section">
      <h3>Kernel Heatmap</h3>
      <div class="chart-container">
        <table class="heatmap">
          <thead>
            <tr>
              <th>Kernel</th>
              ${sortedWeeks.map(w => `<th>${w.slice(5)}</th>`).join('')}
            </tr>
          </thead>
          <tbody>
            ${allKernels.map(k => `
              <tr>
                <td style="font-weight:600">${k}</td>
                ${sortedWeeks.map(w => {
                  const run = weeks[w].find(r => r.kernel === k);
                  if (!run) return '<td><span class="heatmap-cell empty"></span></td>';
                  return `<td><span class="heatmap-cell ${run.display_status}" title="${run.display_status}: ${run.compilation_attempts} compiles"></span></td>`;
                }).join('')}
              </tr>
            `).join('')}
          </tbody>
        </table>
      </div>
    </div>

    ${renderModelComparison(allRuns)}
  `;
}

// ==================== COVERAGE ====================

let coverageData = null;

async function renderCoverage() {
  const content = document.getElementById('content');

  if (!coverageData) {
    content.innerHTML = '<div class="empty-state"><p>Loading coverage data...</p></div>';
    try {
      const resp = await fetch('/api/coverage');
      if (!resp.ok) throw new Error('not found');
      coverageData = await resp.json();
    } catch (e) {
      content.innerHTML = '<div class="empty-state"><h2>Coverage data unavailable</h2><p>Ensure quasar_llk_gap_analysis.xlsx exists and openpyxl is installed.</p></div>';
      return;
    }
  }

  const d = coverageData;
  const summary = d.summary || {};

  // Compute totals
  const sfpuImpl = d.sfpu_kernels.filter(k => k.status === 'Implemented').length;
  const sfpuTotal = d.sfpu_kernels.length;
  const subImpl = d.submodule.filter(s => s.status === 'Implemented' || s.status === 'Quasar-specific').length;
  const subTotal = d.submodule.filter(s => s.status !== 'Quasar-specific').length + d.submodule.filter(s => s.status === 'Quasar-specific').length;
  const apiImpl = d.intermediate_api.filter(a => a.status === 'Implemented').length;
  const apiTotal = d.intermediate_api.length;
  const computeImpl = d.compute_api.reduce((s, f) => s + f.implemented, 0);
  const computeTotal = d.compute_api.reduce((s, f) => s + f.total, 0);

  function pctBar(done, total, label) {
    const pct = total > 0 ? Math.round(done / total * 100) : 0;
    const color = pct >= 80 ? 'var(--green)' : pct >= 40 ? 'var(--yellow)' : 'var(--red)';
    return '<div style="margin-bottom:16px">' +
      '<div style="display:flex;justify-content:space-between;margin-bottom:4px">' +
        '<span style="font-weight:600">' + label + '</span>' +
        '<span style="color:' + color + '">' + done + ' / ' + total + ' (' + pct + '%)</span>' +
      '</div>' +
      '<div style="background:var(--surface);border-radius:4px;height:20px;overflow:hidden;border:1px solid var(--border)">' +
        '<div style="background:' + color + ';height:100%;width:' + pct + '%;transition:width 0.3s"></div>' +
      '</div>' +
    '</div>';
  }

  let coverageView = 'sfpu';
  function renderTab() { renderCoverageDetail(coverageView); }

  content.innerHTML = `
    <div class="detail-panel" style="max-width:1200px">
      <h2 style="margin-bottom:16px">Quasar Implementation Coverage</h2>
      <p style="color:var(--text-muted);margin-bottom:24px">SFPU Kernels &amp; LLK Submodule: live file scan (auto-updates). Intermediate &amp; Compute API: from Excel (tt-metal files, not in this repo).</p>

      ${pctBar(sfpuImpl, sfpuTotal, 'SFPU Kernels')}
      ${pctBar(subImpl, subTotal, 'LLK Submodule Files')}
      ${pctBar(apiImpl, apiTotal, 'Intermediate API Files')}
      ${pctBar(computeImpl, computeTotal, 'Compute API Functions')}

      <div style="display:flex;gap:8px;margin:24px 0 16px">
        <span class="agent-log-btn active" id="cov-tab-sfpu" onclick="setCovTab('sfpu',this)">SFPU Kernels (${sfpuTotal})</span>
        <span class="agent-log-btn" id="cov-tab-submodule" onclick="setCovTab('submodule',this)">LLK Submodule (${subTotal})</span>
        <span class="agent-log-btn" id="cov-tab-api" onclick="setCovTab('api',this)">Intermediate API (${apiTotal})</span>
        <span class="agent-log-btn" id="cov-tab-compute" onclick="setCovTab('compute',this)">Compute API (${d.compute_api.length} files)</span>
      </div>
      <div id="cov-detail"></div>
    </div>
  `;

  setCovTab('sfpu', document.getElementById('cov-tab-sfpu'));
}

function setCovTab(tab, btn) {
  document.querySelectorAll('[id^="cov-tab-"]').forEach(b => b.classList.remove('active'));
  if (btn) btn.classList.add('active');
  renderCoverageDetail(tab);
}

function renderCoverageDetail(tab) {
  const el = document.getElementById('cov-detail');
  if (!el || !coverageData) return;
  const d = coverageData;

  function statusBadge(status) {
    if (status === 'Implemented') return '<span class="status-badge success" style="font-size:11px"><span class="status-dot success"></span>Implemented</span>';
    if (status === 'Quasar-specific') return '<span class="status-badge compiled" style="font-size:11px"><span class="status-dot compiled"></span>QSR-only</span>';
    if (status === 'Not implemented (stub)') return '<span class="status-badge" style="font-size:11px;background:rgba(210,153,34,0.12);color:var(--yellow)"><span class="status-dot" style="background:var(--yellow)"></span>Stub</span>';
    return '<span class="status-badge failed" style="font-size:11px"><span class="status-dot failed"></span>Missing</span>';
  }

  if (tab === 'sfpu') {
    const impl = d.sfpu_kernels.filter(k => k.status === 'Implemented');
    const missing = d.sfpu_kernels.filter(k => k.status !== 'Implemented');
    el.innerHTML = `
      <table class="runs-table">
        <thead><tr><th>Kernel</th><th>BH</th><th>QSR</th><th>Status</th></tr></thead>
        <tbody>
          ${impl.map(k => `<tr><td><strong>${k.kernel}</strong></td><td>${k.bh}</td><td>${k.qsr}</td><td>${statusBadge(k.status)}</td></tr>`).join('')}
          ${missing.map(k => `<tr style="opacity:0.7"><td>${k.kernel}</td><td>${k.bh}</td><td>${k.qsr}</td><td>${statusBadge(k.status)}</td></tr>`).join('')}
        </tbody>
      </table>
    `;
  } else if (tab === 'submodule') {
    const categories = [...new Set(d.submodule.map(s => s.category))];
    el.innerHTML = categories.map(cat => {
      const items = d.submodule.filter(s => s.category === cat);
      const impl = items.filter(s => s.status === 'Implemented' || s.status === 'Quasar-specific').length;
      return '<div style="margin-bottom:16px">' +
        '<h4 style="margin-bottom:8px">' + cat + ' <span style="color:var(--text-muted);font-weight:normal;font-size:12px">' + impl + '/' + items.length + ' implemented</span></h4>' +
        '<table class="runs-table"><thead><tr><th>BH File</th><th>QSR File</th><th>Status</th><th>Notes</th></tr></thead><tbody>' +
        items.map(s =>
          '<tr' + (s.status !== 'Implemented' && s.status !== 'Quasar-specific' ? ' style="opacity:0.7"' : '') + '>' +
          '<td><strong>' + s.bh_file + '</strong></td>' +
          '<td style="font-size:12px">' + s.qsr_file + '</td>' +
          '<td>' + statusBadge(s.status) + '</td>' +
          '<td style="font-size:12px;color:var(--text-muted)">' + (s.notes || '') + '</td></tr>'
        ).join('') +
        '</tbody></table></div>';
    }).join('');
  } else if (tab === 'api') {
    el.innerHTML = `
      <table class="runs-table">
        <thead><tr><th>BH API File</th><th>QSR Equivalent</th><th>Status</th><th>Notes</th></tr></thead>
        <tbody>
          ${d.intermediate_api.map(a =>
            '<tr' + (a.status !== 'Implemented' ? ' style="opacity:0.7"' : '') + '>' +
            '<td><strong>' + a.bh_file + '</strong></td>' +
            '<td style="font-size:12px">' + a.qsr_file + '</td>' +
            '<td>' + statusBadge(a.status) + '</td>' +
            '<td style="font-size:12px;color:var(--text-muted)">' + (a.notes || '') + '</td></tr>'
          ).join('')}
        </tbody>
      </table>
    `;
  } else if (tab === 'compute') {
    el.innerHTML = `
      <table class="runs-table">
        <thead><tr><th>Compute API File</th><th>Implemented</th><th>Stub</th><th>Missing</th><th>Total</th><th>Coverage</th></tr></thead>
        <tbody>
          ${d.compute_api.map(f => {
            const pct = f.total > 0 ? Math.round(f.implemented / f.total * 100) : 0;
            const color = pct >= 80 ? 'var(--green)' : pct >= 40 ? 'var(--yellow)' : f.implemented > 0 ? 'var(--yellow)' : 'var(--red)';
            return '<tr>' +
              '<td><strong>' + f.file + '</strong></td>' +
              '<td style="color:var(--green)">' + f.implemented + '</td>' +
              '<td style="color:var(--yellow)">' + f.stub + '</td>' +
              '<td style="color:var(--red)">' + f.missing + '</td>' +
              '<td>' + f.total + '</td>' +
              '<td>' +
                '<div style="display:flex;align-items:center;gap:8px">' +
                  '<div style="flex:1;background:var(--surface);border-radius:3px;height:12px;overflow:hidden;border:1px solid var(--border);min-width:60px">' +
                    '<div style="background:' + color + ';height:100%;width:' + pct + '%"></div>' +
                  '</div>' +
                  '<span style="color:' + color + ';font-size:12px;font-weight:600;min-width:35px">' + pct + '%</span>' +
                '</div>' +
              '</td></tr>';
          }).join('')}
        </tbody>
      </table>
    `;
  }
}

// ==================== MODEL COMPARISON ====================

function renderModelComparison(runs) {
  const models = [...new Set(runs.map(r => r.model || 'unknown'))].sort();
  if (models.length < 2 && models[0] === 'unknown') {
    return `<div class="trend-section"><h3>Model Comparison</h3><p style="color:var(--text-muted)">No model data yet. Set CODEGEN_MODEL to start tracking.</p></div>`;
  }

  const stats = models.map(m => {
    const mr = runs.filter(r => (r.model || 'unknown') === m);
    const total = mr.length;
    const success = mr.filter(r => r.display_status === 'success').length;
    const compiled = mr.filter(r => r.display_status === 'compiled').length;
    const failed = mr.filter(r => r.display_status === 'failed').length;
    const avgCompiles = total > 0 ? (mr.reduce((s, r) => s + (r.compilation_attempts || 0), 0) / Math.max(mr.reduce((s, r) => s + (r.phases_total || 0), 0), 1)).toFixed(1) : '—';
    const avgDuration = total > 0 ? Math.round(mr.reduce((s, r) => s + (r.duration_seconds || 0), 0) / total) : 0;
    const totalTokens = mr.reduce((s, r) => s + ((r.tokens && r.tokens.total) || 0), 0);
    const avgTokens = total > 0 ? Math.round(totalTokens / total) : 0;
    return { model: m, total, success, compiled, failed, avgCompiles, avgDuration, avgTokens };
  });

  return `
    <div class="trend-section">
      <h3>Model Comparison</h3>
      <div class="chart-container">
        <table class="heatmap" style="text-align:center">
          <thead>
            <tr>
              <th style="text-align:left">Model</th>
              <th>Runs</th>
              <th>Success</th>
              <th>Compiled</th>
              <th>Failed</th>
              <th>Success Rate</th>
              <th>Avg Compiles/Phase</th>
              <th>Avg Duration</th>
              <th>Avg Tokens</th>
            </tr>
          </thead>
          <tbody>
            ${stats.map(s => `
              <tr>
                <td style="font-weight:600;text-align:left">${s.model}</td>
                <td>${s.total}</td>
                <td style="color:var(--green)">${s.success}</td>
                <td style="color:var(--blue)">${s.compiled}</td>
                <td style="color:var(--red)">${s.failed}</td>
                <td style="font-weight:600;color:${s.total > 0 && s.success / s.total >= 0.7 ? 'var(--green)' : s.total > 0 && s.success / s.total >= 0.4 ? 'var(--yellow)' : 'var(--red)'}">${s.total > 0 ? Math.round(s.success / s.total * 100) : 0}%</td>
                <td>${s.avgCompiles}</td>
                <td>${formatDuration(s.avgDuration)}</td>
                <td>${formatTokens(s.avgTokens)}</td>
              </tr>
            `).join('')}
          </tbody>
        </table>
      </div>
    </div>
  `;
}

// ==================== PIPELINE ====================

const AGENT_ORDER = ['agent_arch_lookup', 'agent_analyzer', 'agent_planner', 'agent_writer', 'agent_debugger', 'agent_test_writer', 'agent_tester', 'agent_prettifier'];

function orderAgentLogs(names) {
  const ordered = AGENT_ORDER.filter(a => names.includes(a));
  const extra = names.filter(a => !AGENT_ORDER.includes(a)).sort();
  return ordered.concat(extra);
}

function renderPipeline(run, agentLogs) {
  const agents = run.agents || [];
  const logs = agentLogs || {};
  const phases = run.per_phase || [];

  const ca = run.compilation_attempts || 0;
  const debugCycles = run.debug_cycles || 0;
  const phasesTotal = run.phases_total || 1;
  const tr = run.tests_total || 0;
  const tp = run.tests_passed || 0;
  const hasCompileDebug = ca > phasesTotal;
  const hasTestFailure = tr > 0 && tp < tr;

  const steps = [
    { id: 'arch_lookup', name: 'Research', desc: 'Architecture discovery via Confluence & DeepWiki', log: 'agent_arch_lookup' },
    { id: 'analyzer', name: 'Analyze', desc: 'Study reference implementation, identify phases', log: 'agent_analyzer' },
    { id: 'planner', name: 'Plan', desc: 'Map instructions, design target implementation', log: 'agent_planner' },
    { id: 'writer', name: 'Write', desc: 'Generate kernel code from spec', log: 'agent_writer' },
    { id: 'fix_compile', name: 'Fix Compile', desc: 'Fix compilation errors', log: 'agent_debugger' },
    { id: 'test_writer', name: 'Create Tests', desc: 'Create test suite if none exists', log: 'agent_test_writer' },
    { id: 'tester', name: 'Test', desc: 'Run functional tests on simulator', log: 'agent_tester' },
    { id: 'fix_tests', name: 'Fix Tests', desc: 'Fix runtime/test failures', log: 'agent_debugger' },
    { id: 'prettifier', name: 'Prettify', desc: 'Refactor for readability & style', log: 'agent_prettifier' },
  ];

  return steps.map((step, i) => {
    let ran = false;
    let statusText = '';

    if (step.id === 'fix_compile') {
      ran = hasCompileDebug;
      if (!ran) { statusText = 'not needed'; }
      else { statusText = '⚠ ' + (ca - phasesTotal) + ' extra compile' + (ca - phasesTotal > 1 ? 's' : ''); }
    } else if (step.id === 'test_writer') {
      ran = agents.includes('test_writer') || !!logs['agent_test_writer'];
      if (!ran) { statusText = 'tests existed'; }
      else { statusText = '✓ tests created'; }
    } else if (step.id === 'fix_tests') {
      ran = hasTestFailure && debugCycles > 0;
      if (!ran) { statusText = tr === 0 ? 'no tests to fix' : 'not needed'; }
      else { statusText = '⚠ fixing test failures'; }
    } else if (step.id === 'writer') {
      ran = agents.includes('writer') || !!logs['agent_writer'];
      statusText = ran ? (ca <= phasesTotal ? '✓ clean compile' : '✓ compiled (' + ca + ' attempts)') : 'skipped';
    } else if (step.id === 'tester') {
      ran = agents.includes('tester') || !!logs['agent_tester'];
      if (!ran) { statusText = 'skipped'; }
      else { statusText = tr === 0 ? '— no tests available' : tp === tr ? '✓ ' + tp + '/' + tr + ' passed' : '✗ ' + tp + '/' + tr + ' passed'; }
    } else if (step.id === 'prettifier') {
      ran = agents.includes('prettifier') || !!logs['agent_prettifier'];
      statusText = ran ? (run.prettified ? '✓ done' : '— skipped') : 'skipped';
    } else {
      ran = agents.includes(step.id) || !!logs[step.log];
      statusText = ran ? '✓ done' : 'skipped';
    }

    const nodeClass = ran ? 'ran' : 'skipped';
    const arrow = i < steps.length - 1 ? '<span class="pipeline-arrow">→</span>' : '';

    let node = `<div class="pipeline-node ${nodeClass}">
        <div class="step-num">Step ${i + 1}</div>
        <div class="step-name">${step.name}</div>
        <div class="step-desc">${step.desc}</div>
        <div class="step-status">${statusText}</div>
      </div>`;

    return `<div class="pipeline-step">${node}${arrow}</div>`;
  }).join('');
}

// ==================== UTILS ====================

function formatTokens(n) {
  if (!n || n === 0) return '0';
  if (n >= 1000000) return (n / 1000000).toFixed(1) + 'M';
  if (n >= 1000) return (n / 1000).toFixed(1) + 'k';
  return n.toString();
}

function formatDuration(seconds) {
  if (!seconds) return '—';
  if (seconds < 60) return seconds + 's';
  const m = Math.floor(seconds / 60);
  const s = seconds % 60;
  if (m < 60) return m + 'm ' + s + 's';
  const h = Math.floor(m / 60);
  return h + 'h ' + (m % 60) + 'm';
}

function escapeHtml(str) {
  return str.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
}

// Init
init();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="LLK CodeGen Dashboard")
    parser.add_argument("--port", type=int, default=8080, help="Port to serve on")
    parser.add_argument("--host", type=str, default="0.0.0.0", help="Host to bind to")
    args = parser.parse_args()

    print(f"Dashboard: http://localhost:{args.port}")
    print(f"  Network: http://0.0.0.0:{args.port}")
    print(f"  Projects:")
    for pid, proj in PROJECTS.items():
        print(f"    {pid}: logs={proj['logs_root']}, data={proj['runs_jsonl']}")
    app.run(host=args.host, port=args.port, debug=False)
